from openpyxl import load_workbook
import time
from typing import List, Tuple, Any

class Catalogue:
    _cache_data: List[Tuple[Any, ...]] | None = None
    _cache_time: float = 0

    CATALOGUE_PATH = "./src/exel/constant/справочник.xlsx"
    CACHE_TTL = 28800

    def __init__(self, path: str = CATALOGUE_PATH, ttl: int = CACHE_TTL) -> None:
        self.catalogue_path = path
        self.cache_ttl = ttl

    def _take_sheet(self) -> List[Tuple[Any, ...]]:
        if Catalogue._cache_data is not None and time.time() - Catalogue._cache_time < self.cache_ttl:
            return Catalogue._cache_data

        wb = load_workbook(self.catalogue_path, read_only=True)
        ws = wb.active

        Catalogue._cache_data = list(ws.iter_rows(values_only=True))
        Catalogue._cache_time = time.time()
        wb.close()
        return Catalogue._cache_data

    def get_row_by_article(self, article: str) -> Tuple[Any, ...] | None:
        rows = self._take_sheet()
        for row in rows[1:]:
            if row[0] == article:
                return row
        return None
