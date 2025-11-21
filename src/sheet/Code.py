from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

class Code:
    def __init__(self, path):
        self.path = path
        self.ws = self._take_sheet()

    def _take_sheet(self) -> Worksheet:
        wb = load_workbook(self.path, read_only=True)
        ws = wb.active
        return ws
    
    def get_codes(self):
        result = [row[0] for row in self.ws.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True) if row[0] is not None]
        self.ws.parent.close()
        return result
